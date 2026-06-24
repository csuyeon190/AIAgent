import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load gnb.json from root directory
json_path = r'd:\GNBTEST\gnb.json'
with open(json_path, 'r', encoding='utf-8') as f:
    gnb = json.load(f)

all_items = []

def traverse(item, depth, menu_group, parent=None):
    # Determine type
    item_type = item.get('type')
    if not item_type:
        # If type is not defined, check panelType or classify
        item_type = item.get('panelType') or 'untyped'
        
    title = item.get('title', '')
    gnb_title = item.get('gnb_title', '')
    menu_id = item.get('menuId', '')
    url = item.get('url', '')
    
    current_node = {
        'menu_group': menu_group,
        'depth': depth,
        'type': item_type,
        'title': title,
        'gnb_title': gnb_title,
        'menu_id': menu_id,
        'url': url,
        'parent_id': parent.get('menu_id') if parent else None,
        'parent_title': parent.get('title') if parent else None
    }
    all_items.append(current_node)
    
    # Recursively traverse children
    for child in item.get('items', []):
        traverse(child, depth + 1, menu_group, current_node)

# Traverse all 5 menu groups
for group_name in ['menuGroup1', 'menuGroup2', 'menuGroup3', 'menuGroup4', 'menuGroup5']:
    group_items = gnb.get(group_name, [])
    for root_item in group_items:
        traverse(root_item, 1, group_name)

# 1. Sheet 1: Offering Analysis
# Filter for offering-related types in menuGroup1
# Types are: offering (from panelType), offering_main, offering_cate, offering_sub, and offering_item
# (also including untyped items belonging to the offering tree in menuGroup1)
offering_items = [x for x in all_items if x['menu_group'] == 'menuGroup1']

df_offering = pd.DataFrame(offering_items)
# Rename columns for presentation
df_offering_clean = df_offering[[
    'menu_group', 'type', 'title', 'gnb_title', 'menu_id', 'depth', 'url', 'parent_title', 'parent_id'
]].rename(columns={
    'menu_group': '메뉴 그룹',
    'type': '타입',
    'title': '타이틀',
    'gnb_title': 'GNB 타이틀',
    'menu_id': '메뉴 ID (menuId)',
    'depth': 'Depth (깊이)',
    'url': '이동 URL',
    'parent_title': '상위 메뉴명',
    'parent_id': '상위 메뉴 ID'
})

# 2. Sheet 2: Menu Group Hierarchy
# Include menuGroup1, menuGroup2, menuGroup3, menuGroup4 (and menuGroup5)
hierarchy_rows = []
for item in all_items:
    row = {
        '메뉴 그룹': item['menu_group'],
        '타입': item['type'],
        'Depth': item['depth'],
        'Depth 1': '',
        'Depth 2': '',
        'Depth 3': '',
        'Depth 4': '',
        'Depth 5': '',
        '메뉴 ID': item['menu_id'],
        'URL': item['url']
    }
    # Place title in the appropriate Depth column
    depth_col = f"Depth {item['depth']}"
    if depth_col in row:
        row[depth_col] = item['title']
        if item['depth'] == 1 and item['gnb_title']:
            row[depth_col] = f"{item['title']} ({item['gnb_title']})"
    hierarchy_rows.append(row)

df_hierarchy = pd.DataFrame(hierarchy_rows)

# Save to Excel using openpyxl
excel_path = r'd:\GNBTEST\gnb_analysis.xlsx'
writer = pd.ExcelWriter(excel_path, engine='openpyxl')

df_offering_clean.to_excel(writer, sheet_name='오퍼링 메뉴 분석', index=False)
df_hierarchy.to_excel(writer, sheet_name='전체 메뉴 그룹 계층', index=False)

# Formatting using openpyxl
workbook = writer.book

# Colors & Fonts
header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Navy
data_font = Font(name='맑은 고딕', size=10)
border_side = Side(style='thin', color='D9D9D9')
data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

# Styles for sheets
for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]
    ws.views.sheetView[0].showGridLines = True
    
    # Enable header styling
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = data_border
        
    # Format data cells
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            # Alignments based on content
            header_name = ws.cell(row=1, column=col_idx).value
            if header_name in ['Depth', 'Depth (깊이)', '메뉴 ID', '메뉴 ID (menuId)', '타입', '메뉴 그룹', '상위 메뉴 ID']:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # handle korean characters (approximate width calculation)
            cell_len = 0
            for char in val_str:
                if ord(char) > 127:
                    cell_len += 2
                else:
                    cell_len += 1
            max_len = max(max_len, cell_len)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

writer.close()
print("Excel file created successfully at:", excel_path)
