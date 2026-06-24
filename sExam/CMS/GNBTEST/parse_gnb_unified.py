import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load gnb.json from root directory
json_path = r'd:\GNBTEST\gnb.json'
with open(json_path, 'r', encoding='utf-8') as f:
    gnb = json.load(f)

all_items = []

def traverse(item, depth, menu_group, parent=None):
    # Determine type
    item_type = item.get('type')
    if not item_type:
        item_type = item.get('panelType') or 'untyped'
        
    title = item.get('title', '')
    gnb_title = item.get('gnb_title', '')
    menu_id = item.get('menuId', '')
    url = item.get('url', '')
    
    current_node = {
        'menu_group': menu_group,
        'type': item_type,
        'title': title,
        'menu_id': menu_id,
        'gnb_title': gnb_title,
        'depth': depth,
        'url': url,
        'parent_title': parent.get('title') if parent else '',
        'parent_id': parent.get('menu_id') if parent else ''
    }
    all_items.append(current_node)
    
    # Recursively traverse children
    for child in item.get('items', []):
        traverse(child, depth + 1, menu_group, current_node)

# Traverse all 5 menu groups
for group_name in ['menuGroup1', 'menuGroup2', 'menuGroup3', 'menuGroup4', 'menuGroup5']:
    group_items = gnb.get(group_name, [])
    for root_item in group_items:
        traverse(root_item, 1, group_name)

df_unified = pd.DataFrame(all_items)

# Rename columns according to user request
df_unified_clean = df_unified[[
    'menu_group', 'type', 'title', 'menu_id', 'gnb_title', 'depth', 'url', 'parent_title', 'parent_id'
]].rename(columns={
    'menu_group': '메뉴그룹',
    'type': '타입',
    'title': '타이틀',
    'menu_id': 'menuid',
    'gnb_title': 'gnb 타이틀',
    'depth': 'depth',
    'url': '이동 URL',
    'parent_title': '상위 메뉴명',
    'parent_id': '상위 메뉴ID'
})

# Save to a new Excel file
excel_path = r'd:\GNBTEST\gnb_unified_analysis.xlsx'
writer = pd.ExcelWriter(excel_path, engine='openpyxl')
df_unified_clean.to_excel(writer, sheet_name='GNB 통합 분석', index=False)

# Formatting using openpyxl
workbook = writer.book
ws = workbook['GNB 통합 분석']
ws.views.sheetView[0].showGridLines = True

# Style settings
header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Navy
data_font = Font(name='맑은 고딕', size=10)
border_side = Side(style='thin', color='D9D9D9')
data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

# Enable header styling
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = data_border
    
# Format data cells
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = data_border
        
        # Alignments based on column names
        header_name = ws.cell(row=1, column=col_idx).value
        if header_name in ['메뉴그룹', '타입', 'menuid', 'depth', '상위 메뉴ID']:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Freeze the first row so it stays visible when scrolling
ws.freeze_panes = 'A2'

# Adjust column widths dynamically
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        # Handle Korean characters width
        cell_len = 0
        for char in val_str:
            if ord(char) > 127:
                cell_len += 2
            else:
                cell_len += 1
        max_len = max(max_len, cell_len)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

writer.close()
print("Unified Excel file created successfully at:", excel_path)
